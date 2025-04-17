from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Alignment, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
import os

class ExcelTools:
    @staticmethod
    def get_third_last_row(input_data) -> int:
        """
        获取Excel文件中倒数第三行（不包括空行）的行号
        
        Args:
            input_data: 可以是Excel文件路径(str)或工作表对象(Worksheet)
            
        Returns:
            int: 倒数第三行的行号（从1开始计数）
        """
        # 如果输入是字符串，则加载工作簿
        if isinstance(input_data, str):
            wb = load_workbook(input_data)
            ws = wb.active
        else:
            ws = input_data
        
        # 获取所有非空行的行号
        non_empty_rows = []
        for row in ws.iter_rows():
            # 检查行是否为空（所有单元格都为空）
            if any(cell.value is not None for cell in row):
                non_empty_rows.append(row[0].row)  # 获取行号
        
        if len(non_empty_rows) < 3:
            raise ValueError("Excel文件中的非空行数少于3行")
        
        # 获取倒数第三行的行号
        third_last_row = non_empty_rows[-3]
        
        return third_last_row

    @staticmethod
    def move_rows(ws, start_row, move_steps):
        """
        移动指定行数的数据，包括样式和合并信息
        
        Args:
            ws: openpyxl 的工作表对象
            start_row: 起始行号
            move_steps: 移动的步数（正数表示向下移动，负数表示向上移动）
        """
        try:
            # 获取工作表的列数
            max_col = ws.max_column
            
            # 计算目标行号
            target_row = start_row + move_steps
            
            # 保存源范围的样式和合并信息
            source_range = f"A{start_row}:{get_column_letter(max_col)}{start_row + 2}"
            
            # 先处理合并单元格
            merged_cells_to_process = []
            # 创建合并单元格列表的副本
            merged_ranges = list(ws.merged_cells.ranges)
            for merged_cell in merged_ranges:
                try:
                    # 如果合并单元格在源范围内
                    if merged_cell.min_row >= start_row and merged_cell.max_row <= start_row + 2:
                        merged_cells_to_process.append(merged_cell)
                        # 取消合并
                        ws.unmerge_cells(str(merged_cell))
                except Exception as e:
                    print(f"取消合并单元格时出错: {e}")
                    continue
            
            # 移动数据
            ws.move_range(
                source_range,  # 源范围（三行数据）
                rows=move_steps,  # 移动的步数
                cols=0  # 移动的列数
            )
            
            # 在新的位置重新合并单元格
            for merged_cell in merged_cells_to_process:
                try:
                    new_merged_range = f"{get_column_letter(merged_cell.min_col)}{merged_cell.min_row + move_steps}:{get_column_letter(merged_cell.max_col)}{merged_cell.max_row + move_steps}"
                    ws.merge_cells(new_merged_range)
                except Exception as e:
                    print(f"重新合并单元格时出错: {e}")
                    continue
                
        except Exception as e:
            print(f"移动行时出错: {e}")
            # 如果出错，尝试回滚操作
            try:
                # 重新合并之前取消合并的单元格
                for merged_cell in merged_cells_to_process:
                    try:
                        ws.merge_cells(str(merged_cell))
                    except:
                        continue
            except:
                pass


def init_cell_style(cell, row_height=24, font_bold=False, font_size=12, font_name='新宋体', border_all=True, vertical_alignment='center', horizontal_alignment='center', date_format=None, wrap_text=True):
    """
    设置单元格样式
    
    Args:
        cell: openpyxl 的单元格对象
        row_height: 行高，如果为 None 则保持默认
        font_bold: 是否加粗字体
        font_size: 字体大小，默认为11
        font_name: 字体名称，默认为新宋体
        border_all: 是否添加所有边框
        vertical_alignment: 垂直对齐方式 ('top', 'center', 'bottom')
        horizontal_alignment: 水平对齐方式 ('left', 'center', 'right')
        date_format: 日期格式，例如 'm月d日' 表示 "3月14日" 这样的格式
        wrap_text: 是否自动换行，默认为True
    """
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid") 
    # 设置行高
    if row_height is not None:
        cell.parent.row_dimensions[cell.row].height = row_height
    
    # 设置字体
    cell.font = Font(name=font_name, bold=font_bold, size=font_size)
    
    # 设置边框
    if border_all:
        thin = Side(border_style="thin", color="000000")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    # 设置对齐方式
    cell.alignment = Alignment(
        vertical=vertical_alignment,
        horizontal=horizontal_alignment,
        wrap_text=wrap_text  # 设置自动换行
    )

    # 设置背景颜色
    cell.fill = red_fill


def set_date_format(cell):
    date_format='m月d日'
    if is_column_b(cell):
        cell.number_format = date_format

def is_column_b(cell) -> bool:
    """
    判断单元格是否在B列
    
    Args:
        cell: openpyxl 的单元格对象
        
    Returns:
        bool: 如果单元格在B列返回True，否则返回False
    """
    return cell.column_letter == 'B'  # 或者使用 cell.column == 2
#单独设置字号
def set_font_size(cell, font_size=10, col_list=["E","F","G"]):
    """
    设置指定列的字号
    
    Args:
        cell: openpyxl 的单元格对象
        font_size: 要设置的字体大小
        col_list: 需要设置字号的列列表，默认为 ["E","F","G"]
    """
    if cell.column_letter in col_list:
        cell.font = Font(size=font_size)

#单独设置字体
def set_font_name(cell,font_name = "宋体"):
    if cell.column_letter == "A":
        cell.font = Font(name=font_name,size=12)

#设置数值格式（保留两位小数）
def set_number_format(cell, decimal_places=2):
    """
    设置单元格为数值格式并保留指定小数位数
    
    Args:
        cell: openpyxl 的单元格对象
        decimal_places: 要保留的小数位数，默认为2
    """
    # 设置数值格式，例如 "0.00" 表示保留两位小数
    if cell.column_letter == "I":
        cell.number_format = f"0.{'0' * decimal_places}"
    if cell.column_letter in ["J","K"]:
        cell.number_format = f"0.{'0' * 4}"

def set_thick_right_border(cell):
    """
    设置单元格的粗线右边框，同时保留其他边框
    
    Args:
        cell: openpyxl 的单元格对象
    """
    # 获取当前单元格的边框
    current_border = cell.border
    
    # 创建新的粗线右边框
    thick_right = Side(border_style="medium", color="000000")
    
    # 创建新的边框对象，保持其他边框不变，只修改右边框
    if cell.column_letter == "L":
        new_border = Border(
            left=current_border.left,
            right=thick_right,
            top=current_border.top,
            bottom=current_border.bottom
        )
    
        # 应用新的边框
        cell.border = new_border
        
def set_cell_style(cell):
    init_cell_style(cell)
    set_date_format(cell)
    set_font_size(cell)
    set_font_name(cell)
    set_number_format(cell)
    set_thick_right_border(cell)

def get_start_no(ws) -> int:
    """
    获取包含"客户回签："的单元格的行号
    
    Args:
        ws: openpyxl 的工作表对象
        
    Returns:
        int: 包含"客户回签："的单元格的行号，如果没找到返回0
    """
    for row in ws.iter_rows(min_col=1, max_col=1):  # 只遍历A列
        cell = row[0]
        if cell.value == "客户回签：":  # 找到目标单元格
            # 获取当前行号减3的单元格
            # print(cell.row)
            target_cell = ws.cell(row=cell.row - 3, column=1)
            return target_cell.value if target_cell.value else 0
    
    return 0  # 如果没找到，返回0 

def move_rows(ws, start_row, move_steps):
    """
    移动指定行数的数据，包括样式和合并信息
    
    Args:
        ws: openpyxl 的工作表对象
        start_row: 起始行号
        move_steps: 移动的步数（正数表示向下移动，负数表示向上移动）
    """
    # 获取工作表的列数
    max_col = ws.max_column
    
    # 计算目标行号
    target_row = start_row + move_steps
    
    # 保存源范围的样式和合并信息
    source_range = f"A{start_row}:{get_column_letter(max_col)}{start_row + 2}"
    
    # 先处理合并单元格
    merged_cells_to_process = []
    for merged_cell in ws.merged_cells.ranges:
        # 如果合并单元格在源范围内
        if merged_cell.min_row >= start_row and merged_cell.max_row <= start_row + 2:
            merged_cells_to_process.append(merged_cell)
            # 取消合并
            ws.unmerge_cells(str(merged_cell))
    
    # 移动数据
    ws.move_range(
        source_range,  # 源范围（三行数据）
        rows=move_steps,  # 移动的步数
        cols=0  # 移动的列数
    )
    
    # 在新的位置重新合并单元格
    for merged_cell in merged_cells_to_process:
        try:
            new_merged_range = f"{get_column_letter(merged_cell.min_col)}{merged_cell.min_row + move_steps}:{get_column_letter(merged_cell.max_col)}{merged_cell.max_row + move_steps}"
            ws.merge_cells(new_merged_range)
        except Exception as e:
            print(f"合并单元格时出错: {e}")
            continue 


def sum_k_column_and_save(ws):
    """
    遍历K列，计算所有数值的总和（除了倒数第3行），保留两位小数，并保存到K列倒数第3行
    
    Args:
        ws: openpyxl 的工作表对象
    """
    # 获取K列的最后一行
    last_row = ws.max_row
    
    # 计算倒数第3行的行号
    third_last_row = last_row - 2
    
    total = 0
    # 遍历K列的所有行
    for row in range(1, last_row + 1):
        # 跳过倒数第3行
        if row == third_last_row:
            continue
            
        cell = ws[f"K{row}"]
        # 如果单元格有公式，计算其值
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            # 解析公式
            formula = cell.value[1:]  # 去掉等号
            # 如果是乘法公式
            if '*' in formula:
                cells = formula.split('*')
                if len(cells) == 2:
                    try:
                        # 获取两个单元格的值
                        cell1 = ws[cells[0].strip()]
                        cell2 = ws[cells[1].strip()]
                        if cell1.value is not None and cell2.value is not None:
                            value = float(cell1.value) * float(cell2.value)
                            total += value
                    except (ValueError, TypeError):
                        continue
        # 如果单元格有数值，则累加
        elif cell.value is not None and isinstance(cell.value, (int, float)):
            total += cell.value
    
    # 保留两位小数
    total = round(total, 2)
    
    # 将结果保存到K列倒数第3行
    target_cell = ws[f"K{third_last_row}"]
    target_cell.value = total
    
    # 设置数值格式为保留两位小数
    target_cell.number_format = "0.00"

