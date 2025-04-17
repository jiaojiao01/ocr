import sys
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
from order.order_info import extract_all_info
from tools.commodity_name_handle import split_instrument_name,extract_purchase_order_no,extract_basic_unit
from tools.time_tools import convert_date_format
from tools.excel_tools import ExcelTools,set_cell_style,get_start_no,sum_k_column_and_save
order = extract_all_info(excel_file_path = "../output/table_data_formatted.xlsx")


def handle_headers(order,ws):
    # 使用openpyxl读取excel文件

    # 1.处理A4单元格数据
    ws['A4'] = f"客    户： {order.customer.customer_name:<{62-len(order.customer.customer_name)+len('客    户： ')}}供应商编号:（采购单上有对应的供应商编号）"
    ws['A5'] = f"{'联 系 人：':<77}供方对账联系人: {order.supplier.supplier_contact_person} {order.supplier.supplier_phone}"
    ws['A6'] = f"电    话："
    

def handle_item_list(ws,order,start_row, start_no, move_down_by):
    # 获取第一个工作表
    
    # red_font = Font(color="FF0000")
    for item in order.items:
        start_no += 1
        insert_item = {}
        insert_item["序号"] = start_no
        insert_item["送/退货日期"] = convert_date_format(item.delivery_date)
        insert_item["送/退货单号"] = extract_purchase_order_no(order.purchase_order_no)
        insert_item["采购单号"] = item.requisition_no
        insert_item["料件编号"] = item.material_no
        insert_item["料件名称"] = split_instrument_name(item.product_name)[0]
        insert_item["物料规格"] = split_instrument_name(item.product_name)[1]
        insert_item["单位"] = extract_basic_unit(item.purchase_unit)
        insert_item["数量"] = float(item.purchase_quantity) if item.purchase_quantity else 0
        insert_item["单价（含税）"] = float(item.price_with_tax) if item.price_with_tax else 0
        insert_item["金额（含税）"] = float(item.amount_with_tax) if item.amount_with_tax else 0
        insert_item["备注"] = ""
        
        headers = list(insert_item.keys())
        for col_idx, key in enumerate(headers, start=1):
            # 检查单元格是否被合并
            cell = ws.cell(row=start_row, column=col_idx)
            for merged_range in list(ws.merged_cells.ranges):
                if cell.coordinate in merged_range:
                    # 取消合并
                    ws.unmerge_cells(str(merged_range))
                    break
            
            # 写入值并设置填充
            cell.value = insert_item[key]
            set_cell_style(cell)
            # set_date_format(cell)
        start_row += 1

if __name__ == "__main__":
    excel_path = "../output/对账单格式.xlsx"
    output_path = "../output/对账单格式_new.xlsx"
    wb = load_workbook(excel_path)
    ws = wb.active
    start_no = get_start_no(ws)
    item_count = len(order.items)
    # 处理表头
    handle_headers(order, ws)
    
    # 获取起始行并移动行
    start_row = ExcelTools.get_third_last_row(ws)

    # def handle_item_list(ws,order,start_row, start_no, move_down_by):
    print(f"start_row, start_no + 1, item_count,move_down_by:{start_row, start_no, item_count,item_count}")
    # def move_rows_down(ws,start_row: int, move_down_by: int = 4) -> None:
    #     def move_row_down(ws, target_row):
    ExcelTools.move_rows(ws,start_row + 1 , item_count )
    
    # 处理项目列表
    handle_item_list(ws, order ,start_row + 1, start_no, move_down_by=item_count)
    
    #处理总金额
    sum_k_column_and_save(ws)
    # 保存文件
    wb.save(output_path)
    print(f"文件已保存到: {output_path}")


   
                                       